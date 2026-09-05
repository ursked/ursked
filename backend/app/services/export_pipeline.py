"""
Transformation pipeline for custom data exports.

Before this module the exporter could do `SELECT cols FROM one_source WHERE ...
ORDER BY one_col` and nothing else: no grouping, no aggregates, no column
renaming, no value formatting, no relative date windows. Those are the whole
difference between "a column picker" and "something you would otherwise open
Excel or write SQL for", so they live here rather than being scattered through
the service.

The stages run in a FIXED order, and the order is the contract the UI explains
to the user in plain English:

    1. load        rows come out of the data source
    2. date window absolute dates, or a relative preset resolved against today
    3. filter      row conditions, ANDed
    4. calculate   formula columns (so you can group or filter-after on them)
    5. group       group-by + aggregates, which REPLACES the row shape
    6. sort        one or more keys
    7. limit       top N
    8. project     choose, order, rename and format the output columns

Aggregation deliberately sits after the formula stage: "total overtime per
department" is only expressible if a computed department column already exists.
Sorting sits after aggregation so you can sort by a total.
"""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

# ── Date windows ─────────────────────────────────────────────────────
#
# A saved report that says "last 30 days" has to mean last 30 days *when it
# runs*, not when it was saved. Absolute dates cannot express that, which is why
# every scheduled export currently emails the entire history every time: there
# was nowhere to put a window at all.

DATE_PRESETS = (
    "today",
    "yesterday",
    "last_7_days",
    "last_30_days",
    "last_90_days",
    "this_week",
    "last_week",
    "this_month",
    "last_month",
    "this_quarter",
    "this_year",
    "year_to_date",
    "custom",
)


def resolve_date_window(
    preset: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
    today: Optional[date] = None,
) -> Tuple[Optional[str], Optional[str]]:
    """Turn a preset (or explicit dates) into a concrete ISO from/to pair."""
    if not preset or preset == "custom":
        return date_from, date_to

    d = today or date.today()
    iso = lambda x: x.isoformat()  # noqa: E731

    if preset == "today":
        return iso(d), iso(d)
    if preset == "yesterday":
        y = d - timedelta(days=1)
        return iso(y), iso(y)
    if preset == "last_7_days":
        return iso(d - timedelta(days=6)), iso(d)
    if preset == "last_30_days":
        return iso(d - timedelta(days=29)), iso(d)
    if preset == "last_90_days":
        return iso(d - timedelta(days=89)), iso(d)
    if preset == "this_week":
        start = d - timedelta(days=d.weekday())
        return iso(start), iso(start + timedelta(days=6))
    if preset == "last_week":
        start = d - timedelta(days=d.weekday() + 7)
        return iso(start), iso(start + timedelta(days=6))
    if preset == "this_month":
        start = d.replace(day=1)
        return iso(start), iso(_end_of_month(start))
    if preset == "last_month":
        first_this = d.replace(day=1)
        end_prev = first_this - timedelta(days=1)
        return iso(end_prev.replace(day=1)), iso(end_prev)
    if preset == "this_quarter":
        q_start_month = 3 * ((d.month - 1) // 3) + 1
        start = d.replace(month=q_start_month, day=1)
        end_month = q_start_month + 2
        return iso(start), iso(_end_of_month(d.replace(month=end_month, day=1)))
    if preset == "this_year":
        return iso(d.replace(month=1, day=1)), iso(d.replace(month=12, day=31))
    if preset == "year_to_date":
        return iso(d.replace(month=1, day=1)), iso(d)

    # Unknown preset: fall back to whatever explicit dates were given rather
    # than silently exporting everything.
    return date_from, date_to


def _end_of_month(d: date) -> date:
    if d.month == 12:
        return d.replace(day=31)
    return d.replace(month=d.month + 1, day=1) - timedelta(days=1)


# ── Filters ──────────────────────────────────────────────────────────

FILTER_OPERATORS = {
    "eq": "is",
    "neq": "is not",
    "contains": "contains",
    "not_contains": "does not contain",
    "starts_with": "starts with",
    "ends_with": "ends with",
    "gt": "is greater than",
    "gte": "is greater than or equal to",
    "lt": "is less than",
    "lte": "is less than or equal to",
    "between": "is between",
    "in": "is any of",
    "not_in": "is none of",
    "is_empty": "is empty",
    "is_not_empty": "is not empty",
}

# Operators that carry no value, so the UI hides the value box entirely.
NO_VALUE_OPERATORS = {"is_empty", "is_not_empty"}


def _as_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def match_filter(cell: Any, operator: str, value: Any) -> bool:
    """Evaluate one condition against one cell.

    An UNKNOWN operator raises rather than returning True. The previous
    behaviour fell through to `return True`, so a typo'd operator matched every
    row and the export silently ignored the filter.
    """
    cv = "" if cell is None else str(cell)
    cv_l = cv.lower()

    if operator == "is_empty":
        return cv.strip() == ""
    if operator == "is_not_empty":
        return cv.strip() != ""

    if operator == "in" or operator == "not_in":
        wanted = value if isinstance(value, list) else [
            v.strip() for v in str(value).split(",") if v.strip()
        ]
        hit = cv_l in {str(w).lower() for w in wanted}
        return hit if operator == "in" else not hit

    if operator == "between":
        lo_raw, hi_raw = (value + [None, None])[:2] if isinstance(value, list) else (None, None)
        lo_n, hi_n, cv_n = _as_number(lo_raw), _as_number(hi_raw), _as_number(cell)
        if lo_n is not None and hi_n is not None and cv_n is not None:
            return lo_n <= cv_n <= hi_n
        # Dates and anything else compare lexically, which is correct for ISO.
        lo_s = "" if lo_raw is None else str(lo_raw)
        hi_s = "" if hi_raw is None else str(hi_raw)
        return (not lo_s or cv >= lo_s) and (not hi_s or cv <= hi_s)

    fv = "" if value is None else str(value)
    fv_l = fv.lower()

    if operator == "eq":
        return cv_l == fv_l
    if operator == "neq":
        return cv_l != fv_l
    if operator == "contains":
        return fv_l in cv_l
    if operator == "not_contains":
        return fv_l not in cv_l
    if operator == "starts_with":
        return cv_l.startswith(fv_l)
    if operator == "ends_with":
        return cv_l.endswith(fv_l)

    if operator in ("gt", "gte", "lt", "lte"):
        cv_n, fv_n = _as_number(cell), _as_number(value)
        if cv_n is None or fv_n is None:
            # Dates arrive as ISO strings, which order correctly as text.
            if operator == "gt":
                return cv > fv
            if operator == "gte":
                return cv >= fv
            if operator == "lt":
                return cv < fv
            return cv <= fv
        if operator == "gt":
            return cv_n > fv_n
        if operator == "gte":
            return cv_n >= fv_n
        if operator == "lt":
            return cv_n < fv_n
        return cv_n <= fv_n

    raise ValueError(f"Unknown filter operator: {operator}")


def apply_filters(
    rows: List[Dict[str, Any]], filters: Optional[List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    if not filters:
        return rows
    out = rows
    for f in filters:
        col = f.get("column", "")
        op = f.get("operator", "eq")
        val = f.get("value")
        out = [r for r in out if match_filter(r.get(col, ""), op, val)]
    return out


# ── Aggregation ──────────────────────────────────────────────────────

AGGREGATE_FUNCTIONS = {
    "sum": "Total",
    "avg": "Average",
    "min": "Lowest",
    "max": "Highest",
    "count": "Count",
    "count_distinct": "Distinct count",
    "first": "First",
}


def aggregate_label(func: str, column_label: str) -> str:
    """Human label for a generated column, e.g. "Total Overtime (min)"."""
    if func == "count":
        return "Number of rows"
    return f"{AGGREGATE_FUNCTIONS.get(func, func.title())} {column_label}"


def apply_grouping(
    rows: List[Dict[str, Any]],
    group_by: List[str],
    aggregations: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Collapse rows to one per distinct combination of the group-by columns.

    Grouping REPLACES the row shape: the result has exactly the group-by columns
    plus one column per aggregate, keyed by each aggregate's output name. Any
    other selected column is dropped, because there is no defensible value to
    show for it once fifty rows have become one.
    """
    if not group_by:
        return rows

    buckets: Dict[Tuple, List[Dict[str, Any]]] = {}
    order: List[Tuple] = []
    for r in rows:
        key = tuple(str(r.get(g, "")) for g in group_by)
        if key not in buckets:
            buckets[key] = []
            order.append(key)
        buckets[key].append(r)

    out: List[Dict[str, Any]] = []
    for key in order:
        members = buckets[key]
        row: Dict[str, Any] = {g: key[i] for i, g in enumerate(group_by)}
        for agg in aggregations or []:
            out_key = agg.get("output_key") or agg.get("label") or agg.get("column")
            row[out_key] = _compute_aggregate(members, agg)
        out.append(row)
    return out


def _compute_aggregate(members: List[Dict[str, Any]], agg: Dict[str, Any]) -> Any:
    func = agg.get("func", "sum")
    col = agg.get("column", "")

    if func == "count":
        return len(members)

    values = [m.get(col) for m in members]

    if func == "count_distinct":
        return len({str(v) for v in values if v not in (None, "")})
    if func == "first":
        for v in values:
            if v not in (None, ""):
                return v
        return ""

    numbers = [n for n in (_as_number(v) for v in values) if n is not None]
    if not numbers:
        # min/max still make sense on text (earliest date, first name).
        if func in ("min", "max"):
            texts = sorted(str(v) for v in values if v not in (None, ""))
            if not texts:
                return ""
            return texts[0] if func == "min" else texts[-1]
        return 0

    if func == "sum":
        total = sum(numbers)
    elif func == "avg":
        total = sum(numbers) / len(numbers)
    elif func == "min":
        total = min(numbers)
    elif func == "max":
        total = max(numbers)
    else:
        return ""

    # Keep integers looking like integers: a count of shifts should read 12, not
    # 12.0, in a spreadsheet cell people will eyeball.
    rounded = round(total, 4)
    return int(rounded) if float(rounded).is_integer() else rounded


# ── Sorting ──────────────────────────────────────────────────────────


def _sort_key(value: Any):
    if value is None or value == "":
        return (1, 0.0, "")
    n = _as_number(value)
    if n is not None:
        return (0, n, "")
    return (0, 0.0, str(value).lower())


def apply_sort(rows: List[Dict[str, Any]], sorts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Sort by one or more keys.

    Applied last-key-first because Python's sort is stable, which is what makes
    "by department, then by name within it" come out right.
    """
    if not sorts:
        return rows
    out = list(rows)
    for s in reversed(sorts):
        col = s.get("column")
        if not col:
            continue
        desc = str(s.get("direction", "asc")).lower() == "desc"
        out.sort(key=lambda r: _sort_key(r.get(col, "")), reverse=desc)
    return out


# ── Value formatting ─────────────────────────────────────────────────

DATE_PATTERNS = {
    "iso": "%Y-%m-%d",
    "dmy": "%d/%m/%Y",
    "mdy": "%m/%d/%Y",
    "long": "%d %B %Y",
    "month_year": "%B %Y",
    "day_month": "%d %b",
    "weekday": "%a %d %b",
}

TIME_PATTERNS = {"24h": "%H:%M", "12h": "%I:%M %p"}


def format_value(value: Any, spec: Optional[Dict[str, Any]]) -> Any:
    """Apply a display format. Unparseable values are returned untouched.

    Deliberately forgiving: a formatting choice must never turn a cell into an
    error. If a date will not parse, the original string is better than "#ERROR".
    """
    if not spec or value in (None, ""):
        return value

    kind = spec.get("kind")

    if kind == "number":
        n = _as_number(value)
        if n is None:
            return value
        decimals = int(spec.get("decimals", 0) or 0)
        s = f"{n:,.{decimals}f}" if spec.get("thousands") else f"{n:.{decimals}f}"
        prefix = spec.get("prefix") or ""
        suffix = spec.get("suffix") or ""
        return f"{prefix}{s}{suffix}"

    if kind == "date":
        fmt = DATE_PATTERNS.get(spec.get("pattern", "iso"))
        if not fmt:
            return value
        parsed = _parse_date(value)
        return parsed.strftime(fmt) if parsed else value

    if kind == "time":
        fmt = TIME_PATTERNS.get(spec.get("pattern", "24h"))
        if not fmt:
            return value
        parsed = _parse_time(value)
        return parsed.strftime(fmt).lstrip("0") if parsed and spec.get("pattern") == "12h" else (
            parsed.strftime(fmt) if parsed else value
        )

    if kind == "text":
        t = spec.get("transform")
        s = str(value)
        if t == "upper":
            return s.upper()
        if t == "lower":
            return s.lower()
        if t == "title":
            return s.title()
        return s

    return value


def _parse_date(value: Any) -> Optional[date]:
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[: len(fmt) + 2] if "T" in fmt or " " in fmt else s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return None


def _parse_time(value: Any) -> Optional[datetime]:
    s = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ── Output shaping ───────────────────────────────────────────────────


def build_output_columns(
    *,
    columns: List[str],
    custom_columns: Optional[List[Dict[str, str]]],
    group_by: Optional[List[str]],
    aggregations: Optional[List[Dict[str, Any]]],
    label_for: Callable[[str], str],
    aliases: Optional[Dict[str, str]] = None,
) -> List[Tuple[str, str]]:
    """Return the ordered (field_key, header) pairs for the output.

    When grouping is on, the selected columns are irrelevant — the shape is
    group-by keys plus aggregates — so the caller does not have to keep the two
    in sync and the UI can say so plainly.
    """
    aliases = aliases or {}
    pairs: List[Tuple[str, str]] = []

    if group_by:
        for g in group_by:
            pairs.append((g, aliases.get(g) or label_for(g)))
        for agg in aggregations or []:
            key = agg.get("output_key") or agg.get("label") or agg.get("column")
            header = aliases.get(key) or agg.get("label") or aggregate_label(
                agg.get("func", "sum"), label_for(agg.get("column", ""))
            )
            pairs.append((key, header))
        return pairs

    for c in columns:
        pairs.append((c, aliases.get(c) or label_for(c)))
    for cc in custom_columns or []:
        name = cc["name"]
        pairs.append((name, aliases.get(name) or name))
    return pairs


def project_rows(
    rows: List[Dict[str, Any]],
    output_columns: List[Tuple[str, str]],
    formats: Optional[Dict[str, Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    """Reduce each row to the output fields, applying per-column formatting."""
    formats = formats or {}
    out = []
    for r in rows:
        row = {}
        for key, _header in output_columns:
            row[key] = format_value(r.get(key, ""), formats.get(key))
        out.append(row)
    return out


# ── Serialisation ────────────────────────────────────────────────────

# Characters that make a spreadsheet treat a cell as a formula. Employee-entered
# free text flows into exports, so "=cmd|'...'" would execute on open.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def csv_safe(value: Any) -> str:
    if value is None:
        return ""
    s = value if isinstance(value, str) else str(value)
    if s and s[0] in _FORMULA_TRIGGERS:
        return "'" + s
    return s


def generate_csv(rows: List[Dict[str, Any]], output_columns: List[Tuple[str, str]]) -> str:
    import csv as _csv

    buf = io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow([h for _k, h in output_columns])
    for r in rows:
        writer.writerow([csv_safe(r.get(k, "")) for k, _h in output_columns])
    return buf.getvalue()


def generate_xlsx(
    rows: List[Dict[str, Any]],
    output_columns: List[Tuple[str, str]],
    sheet_name: str = "Export",
) -> bytes:
    """Write a real workbook: typed cells, a frozen header, sized columns.

    Numbers are written as numbers and dates as dates, so the recipient can
    sort, sum and pivot without re-typing the whole sheet — which is the main
    reason people ask for Excel instead of CSV in the first place.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    # Excel rejects sheet names over 31 chars or containing []:*?/\
    safe = "".join(c for c in sheet_name if c not in "[]:*?/\\")[:31] or "Export"
    ws.title = safe

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="7C3AED")
    for ci, (_key, header) in enumerate(output_columns, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for ri, r in enumerate(rows, start=2):
        for ci, (key, _header) in enumerate(output_columns, start=1):
            ws.cell(row=ri, column=ci, value=_xlsx_value(r.get(key, "")))

    ws.freeze_panes = "A2"
    for ci, (key, header) in enumerate(output_columns, start=1):
        widest = len(str(header))
        for r in rows[:200]:
            widest = max(widest, len(str(r.get(key, ""))))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(widest + 2, 10), 45)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_value(value: Any) -> Any:
    """Type a cell so Excel treats it as a number or date, not text."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value)
    # A leading apostrophe is our CSV-injection guard; Excel does not need it
    # because openpyxl writes values, not formulas.
    if s and s[0] in _FORMULA_TRIGGERS and not _looks_numeric(s):
        return s
    n = _as_number(s)
    if n is not None and _looks_numeric(s):
        return int(n) if float(n).is_integer() and "." not in s else n
    d = _parse_date(s) if len(s) >= 8 and s[:4].isdigit() else None
    if d:
        return d
    return s


def _looks_numeric(s: str) -> bool:
    t = s.strip().replace(",", "")
    if not t or t in ("-", "+", "."):
        return False
    if t[0] in "+-":
        t = t[1:]
    return t.replace(".", "", 1).isdigit()

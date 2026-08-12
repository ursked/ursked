"""Formatted XLSX work-schedule export.

Reproduces the formal "Regular Work Schedule" layout used by clients: a merged
title, a two-row banded header, and one row per employee per day with Excel
serial dates and time fractions, plus a REMARKS column driven by each leave
type's configurable export_code.

Layout (columns A..K):
    A  EMPLOYEE (formal LASTNAME, First Middle)
    B  WORK SCHEDULE (Dates) / FROM
    C  WORK SCHEDULE (Dates) / TO
    D  DWS  (day work status: "FREE" for rest days, blank otherwise)
    E  WORK SCHEDULE (TIME) / START
    F  WORK SCHEDULE (TIME) / END
    G  1 HR UNPAID BREAK / START
    H  1 HR UNPAID BREAK / END
    I  30 MIN PAID BREAK / START
    J  30 MIN PAID BREAK / END
    K  REMARKS
"""
from __future__ import annotations

import io
from datetime import date, datetime, time, timedelta
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.configurable_types import ScheduleFormat
from app.models.leave import LeaveApplication, LeaveType
from app.models.schedule import DateRemark, Shift
from app.models.user import User

# Excel 1900 date system epoch (accounts for the historical leap-year bug).
_EXCEL_EPOCH = datetime(1899, 12, 30)

REST_STATUSES = {"rest_day", "rest day", "restday", "day_off", "day off", "free"}

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
TITLE_FONT = Font(bold=True, size=12)
HEADER_FONT = Font(bold=True, size=9)
CELL_FONT = Font(size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def _excel_serial(d: date) -> int:
    return (datetime(d.year, d.month, d.day) - _EXCEL_EPOCH).days


def _time_fraction(t: time) -> float:
    return round((t.hour * 3600 + t.minute * 60 + t.second) / 86400.0, 10)


def _add_minutes(t: time, minutes: int) -> time:
    base = datetime(2000, 1, 1, t.hour, t.minute, t.second)
    return (base + timedelta(minutes=minutes)).time()


class ScheduleExportService:

    @staticmethod
    async def _leave_code_map(db: AsyncSession, tenant_id: UUID) -> dict:
        """{leave_type_code: export_code} for mapping a leave day to a remark."""
        rows = (await db.execute(
            select(LeaveType).where(LeaveType.tenant_id == tenant_id)
        )).scalars().all()
        return {lt.code: (lt.export_code or lt.code.upper()) for lt in rows}

    @staticmethod
    async def _schedule_formats(db: AsyncSession, tenant_id: UUID) -> dict:
        rows = (await db.execute(
            select(ScheduleFormat).where(ScheduleFormat.tenant_id == tenant_id)
        )).scalars().all()
        return {f.code: f for f in rows}

    @staticmethod
    async def _holidays(db: AsyncSession, tenant_id: UUID, start: date, end: date) -> set:
        rows = (await db.execute(
            select(DateRemark.date).where(
                DateRemark.tenant_id == tenant_id,
                DateRemark.is_holiday == True,  # noqa: E712
                DateRemark.date >= start,
                DateRemark.date <= end,
            )
        )).all()
        return {r[0] for r in rows}

    @staticmethod
    def _breaks(fmt: Optional[ScheduleFormat], start: time):
        """Derive (paid_start, paid_end, unpaid_start, unpaid_end) from the
        employee's schedule format, relative to the shift start."""
        paid_s = paid_e = unpaid_s = unpaid_e = None
        if fmt and start:
            if (fmt.paid_break_minutes or 0) > 0:
                offset = int((fmt.paid_break_after_hours or 0) * 60)
                paid_s = _add_minutes(start, offset)
                paid_e = _add_minutes(paid_s, int(fmt.paid_break_minutes))
            if (fmt.unpaid_break_minutes or 0) > 0:
                offset = int((fmt.unpaid_break_after_hours or 0) * 60)
                unpaid_s = _add_minutes(start, offset)
                unpaid_e = _add_minutes(unpaid_s, int(fmt.unpaid_break_minutes))
        return paid_s, paid_e, unpaid_s, unpaid_e

    @staticmethod
    async def build_workbook(
        db: AsyncSession,
        tenant_id: UUID,
        start_date: date,
        end_date: date,
        *,
        title: str = "Regular Work Schedule",
        sheet_name: Optional[str] = None,
    ) -> bytes:
        leave_codes = await ScheduleExportService._leave_code_map(db, tenant_id)
        formats = await ScheduleExportService._schedule_formats(db, tenant_id)
        holidays = await ScheduleExportService._holidays(db, tenant_id, start_date, end_date)

        # Employees who have at least one shift in the range, plus their shifts.
        emp_ids = (await db.execute(
            select(Shift.employee_id).where(
                Shift.tenant_id == tenant_id,
                Shift.date >= start_date,
                Shift.date <= end_date,
            ).distinct()
        )).scalars().all()

        employees = (await db.execute(
            select(User).where(User.id.in_(emp_ids))
            .order_by(User.last_name, User.first_name)
        )).scalars().all() if emp_ids else []

        shifts = (await db.execute(
            select(Shift)
            .options(selectinload(Shift.leave_application))
            .where(
                Shift.tenant_id == tenant_id,
                Shift.date >= start_date,
                Shift.date <= end_date,
            )
            .order_by(Shift.employee_id, Shift.date, Shift.sequence_number)
        )).scalars().all()

        # Index shifts by (employee_id, date) — first shift of the day wins.
        by_emp_date: dict = {}
        for s in shifts:
            by_emp_date.setdefault((s.employee_id, s.date), s)

        # All calendar days in range.
        days = []
        d = start_date
        while d <= end_date:
            days.append(d)
            d += timedelta(days=1)

        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_name or f"{start_date:%b %d} - {end_date:%d}")[:31]

        # ── Title (row 1, merged A..J) ──
        ws.merge_cells("A1:J1")
        c = ws["A1"]
        c.value = title
        c.font = TITLE_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")

        # ── Two-row banded header (rows 2-3) ──
        ws.merge_cells("A2:A3")
        ws["A2"] = "EMPLOYEE"
        ws.merge_cells("B2:C2")
        ws["B2"] = "WORK SCHEDULE (Dates)"
        ws.merge_cells("D2:D3")
        ws["D2"] = "DWS"
        ws.merge_cells("E2:F2")
        ws["E2"] = "WORK SCHEDULE\n(TIME)"
        ws.merge_cells("G2:H2")
        ws["G2"] = "1 HR UNPAID BREAK\n(9-HOUR SHIFT)"
        ws.merge_cells("I2:J2")
        ws["I2"] = "30 MIN PAID BREAK\n(8-HOUR SHIFT)"
        ws.merge_cells("K2:K3")
        ws["K2"] = "REMARKS"

        sub = {"B3": "FROM", "C3": "TO", "E3": "START", "F3": "END",
               "G3": "START", "H3": "END", "I3": "START", "J3": "END"}
        for ref, val in sub.items():
            ws[ref] = val

        for row in (2, 3):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.font = HEADER_FONT
                cell.alignment = CENTER
                cell.fill = HEADER_FILL
                cell.border = BORDER

        # ── Data rows ──
        r = 4
        for emp in employees:
            fmt = formats.get(emp.schedule_format)
            for day in days:
                s = by_emp_date.get((emp.id, day))
                dws = ""
                start_t = end_t = None
                remark = ""
                if s is not None:
                    status = (s.status or "").lower()
                    is_rest = status in REST_STATUSES
                    if is_rest:
                        dws = "FREE"
                    else:
                        start_t = s.start_time
                        end_t = s.end_time
                    # Remark: leave code, else holiday (worked days only), else
                    # the shift's own remark text.
                    if s.leave_application is not None:
                        remark = leave_codes.get(s.leave_application.leave_type, "")
                    if not remark and not is_rest and day in holidays:
                        remark = "HOL OFF"
                    if not remark and s.remarks:
                        remark = s.remarks
                else:
                    # No shift row at all → treat as rest/unscheduled.
                    dws = "FREE"

                paid_s = paid_e = unpaid_s = unpaid_e = None
                if start_t:
                    paid_s, paid_e, unpaid_s, unpaid_e = ScheduleExportService._breaks(fmt, start_t)

                ScheduleExportService._write_row(
                    ws, r, emp.formal_name, day, day, dws,
                    start_t, end_t, unpaid_s, unpaid_e, paid_s, paid_e, remark,
                )
                r += 1

        ScheduleExportService._set_widths(ws)
        ws.freeze_panes = "A4"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _write_row(ws, r, name, d_from, d_to, dws, start_t, end_t,
                   unpaid_s, unpaid_e, paid_s, paid_e, remark):
        ws.cell(row=r, column=1, value=name).alignment = LEFT

        def put_date(col, dt):
            cell = ws.cell(row=r, column=col)
            if dt:
                cell.value = _excel_serial(dt)
                cell.number_format = "mm/dd/yyyy"
            cell.alignment = CENTER

        def put_time(col, t):
            cell = ws.cell(row=r, column=col)
            if t:
                cell.value = _time_fraction(t)
                cell.number_format = "h:mm AM/PM"
            cell.alignment = CENTER

        put_date(2, d_from)
        put_date(3, d_to)
        ws.cell(row=r, column=4, value=dws or None).alignment = CENTER
        put_time(5, start_t)
        put_time(6, end_t)
        put_time(7, unpaid_s)
        put_time(8, unpaid_e)
        put_time(9, paid_s)
        put_time(10, paid_e)
        ws.cell(row=r, column=11, value=remark or None).alignment = CENTER

        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.font = CELL_FONT
            cell.border = BORDER

    @staticmethod
    def _set_widths(ws):
        widths = {1: 30, 2: 12, 3: 12, 4: 8, 5: 11, 6: 11,
                  7: 11, 8: 11, 9: 11, 10: 11, 11: 12}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 16
